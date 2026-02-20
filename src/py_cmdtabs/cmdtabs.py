import sys, gzip, os, warnings, copy
import os.path
from collections import defaultdict

class CmdTabs:
	transposed = False
	compressed_input = False
	compressed_output = False

	def load_input_data(input_path, sep="\t", limit=-1, first_only=False, add_empty_fields=True, fill_character = "", autodetect_compression=False):
		is_compressed = CmdTabs.compressed_input
		if autodetect_compression: is_compressed = input_path.endswith('.gz')
		
		open_file = gzip.open if is_compressed else open
		if limit > 0: # THis is due to ruby compute de cuts in other way and this fix enables ruby mode. Think if adapt to python way
			limit -= 1
		if input_path == '-':
			if is_compressed: 
				input_data = gzip.decompress(sys.stdin.buffer.read()).decode().strip().split('\n')
			else: 
				input_data = sys.stdin
		else:
			with open_file(input_path, "rt") as f:
				input_data = f.readlines()
		input_data_arr = []
		for line in input_data:
			fields_number = line.count(sep) + 1
			#print('---', file=sys.stderr)
			#print(repr(), file=sys.stderr)
			fields = line.rstrip().split(sep, limit)
			fields = [ fill_character if field == "" else field for field in fields]
			if add_empty_fields:
				limit_fields = 0 if limit <= 0 else limit 
				fields = fields + ( [fill_character] * (fields_number - len(fields) - limit_fields) )
			input_data_arr.append(fields)
			if first_only:
				break
		if CmdTabs.transposed:
			input_data_arr = CmdTabs.transpose(input_data_arr)

		return input_data_arr

	def load_several_files(all_files, sep = "\t", limit=-1, dict_keys_mapper = None, add_empty_fields=True, autodetect_compression=False):
		loaded_files = {}
		for file in all_files:
			if os.path.isdir(file):
				warnings.warn(file +" is not a valid file")
				continue
			key_id = file if not dict_keys_mapper else dict_keys_mapper(file)
			if autodetect_compression: CmdTabs.compressed_input = file.endswith('.gz') or file.endswith('.gzip') or False
			loaded_files[key_id] = CmdTabs.load_input_data(file, sep, limit, add_empty_fields=add_empty_fields)
		return loaded_files

	def load_files(files_path, fill_character = '-'): #NO TEST # Cleaning an filling behaviour could be sent to load_input_data as options and use on load_several_files
		files = {}
		for file_name in files_path:
			input_table = CmdTabs.load_input_data(file_name)
			file = []
			for fields in input_table:
				if fields.count('') == len(fields): continue #skip blank records
				file.append([ fill_character if field == "" else field for field in fields])
			files[file_name] = file
		return files

	def build_pattern(col_filter, keywords):
		pattern = defaultdict(lambda: False	)
		if col_filter != None and keywords != None:
			if isinstance(keywords, str): # search keywords from given string as species the % and & characters. % is delimiter of keywrod per column and & us delimiter of keywrods in th same column
				keys_per_col = keywords.split('%')
				if len(keys_per_col) != len(col_filter): os.abort('Number of keywords not equal to number of filtering columns') 
				for i, col in enumerate(col_filter): pattern[col] = keys_per_col[i].split('&')
			elif isinstance(keywords, list): # search keywords from given list (file) in any column
				for i, col in enumerate(col_filter): pattern[col] = keywords
		return pattern

	def index_array(table, col_from=0, col_to=1, header=False):
		index = defaultdict(lambda: False)
		if type(col_to) != list:
			f = lambda x: x[col_to]
		else: 
			f = lambda x: [x[c] for c in col_to]
		if header: 
			head = table.pop(0)
			index['header'] = f(head)
		for row in table:
			index[row[col_from]] = f(row)
		return index

	def index_metrics(input_data, attributes):
		n_attrib = len(attributes)
		indexed_metrics = defaultdict(lambda: False	)
		metric_names = []
		for entry in input_data:
			entry = entry.copy()
			sample_id = entry.pop(0)
			sample_attributes =  [entry.pop(0) for i in range(n_attrib)]
			if len(entry) == 2: 
				metric_name, metric = entry # The metric is full and it has key and val
			else:
				metric_name = entry[0] # The metric is corrupted and it has only the key but val
				metric = None

			if metric_name not in metric_names: metric_names.append(metric_name)
			query = indexed_metrics[sample_id]
			if not query:
				indexed_metrics[sample_id] = {metric_name: metric}
				i = 0
				for attrib in attributes:
					indexed_metrics[sample_id][attrib] = sample_attributes[i] 
					i += 1
			else:
				query[metric_name] = metric
		return metric_names, indexed_metrics

	def parse_column_indices(col_str_idxs, has_header= False, table=None):
		if has_header: 
			col_str_idxs = CmdTabs.get_name_to_index_equivalences(col_str_idxs, table[0], idx_offset=1)
	
		cols = []
		for col in col_str_idxs:
			if "-" in col: # Range of columns (both ends included)
				start, end = col.split("-")
				range_cols = [i - 1 for i in range(int(start), int(end) + 1)]
				cols.extend(range_cols)
			else: # Single column
				cols.append(int(col) - 1)
		return cols
	
	def get_name_to_index_equivalences(string_list_to_convert, referece_string_list, idx_offset=1):
		cols_header_dict = dict([(col_name, str(idx+idx_offset)) for idx, col_name in enumerate(referece_string_list)])
		cols_to_get_processed = []
		try:
			for col in string_list_to_convert:
				if col.isdigit(): #Single numeric column case: If the column specifier is a number, we leave as it is
					cols_to_get_processed.append(col)
				elif "-" in col and not "%-%" in col: #Range of numeric column case or named column with hyphen
					start_col, end_col = col.split("-")
					if start_col.isdigit() and end_col.isdigit(): #We check if the column specifier is a range of numeric columns, if so we leave as it is
						cols_to_get_processed.append(col)
					else: #This is an edge of a named column with a hyphen in the name but not numeric range, we convert to number
						cols_to_get_processed.append(cols_header_dict[col])
				elif "%-%" not in col: # Single named column case: This is the case of a single column name, we convert to number
					cols_to_get_processed.append(cols_header_dict[col])
				else: # Range of named columns case: This is the case of a range of column names, we convert to numbers and keep the range format
					start_col, end_col = col.split("%-%")
					start_col = cols_header_dict[start_col]
					end_col = cols_header_dict[end_col]
					cols_to_get_processed.append(f"{start_col}-{end_col}")
		except KeyError as e:
			raise KeyError(f"Column '{e.args[0]}' not found in header. Available columns: {', '.join(cols_header_dict.keys())}")
		return cols_to_get_processed		

	def load_and_parse_tags(tags, sep):
		parsed_tags = []
		for tag in tags:
			if os.path.exists(tag):
				parsed_tags.extend(CmdTabs.load_input_data(tag, sep, first_only=True))
			else:
				parsed_tags.append(tag.split(sep))
		return [item for sublist in parsed_tags for item in sublist] #Flatten list

	def load_records(input_file, cols, full): # TODO: Convert list to string to be hasheable is a workaround. We must search another way to do this!!!
		records = {}
		full_row_of_records = {}
		for fields in input_file:
			field = "\t".join([fields[c] for c in cols]) # Lists are not hasheable in Python so we covert them to string
			records[field] = True
			if full: full_row_of_records[field] = ["\t".join(fields)]
		return list(records.keys()), full_row_of_records


	def aggregate_column(input_table, col_index, cols_agg, sep, agg_mode="concatenate"):
		import py_exp_calc.exp_calc as pxc
		aggregated_data = defaultdict(lambda: False	)
		if type(cols_agg) == int: cols_agg = [cols_agg]
		if type(col_index) == int: col_index = [col_index]
		for fields in input_table:
			for col_agg in cols_agg:
				pxc.add_nested_value(aggregated_data, (tuple([fields[idx] for idx in col_index]), col_agg), fields[col_agg], True)
		aggregated_data_arr = []
		for k, agg_dict in aggregated_data.items():
			col_data = [idx_col for idx_col in k]
			for aggregated_column in agg_dict.values():
				col_data.append( str(CmdTabs._make_one_or_several_aggregation(aggregated_column, agg_mode, sep)) )
			aggregated_data_arr.append(col_data)
		return aggregated_data_arr
	
	def _make_one_or_several_aggregation(aggregated_column, aggregation_modes, sep):
		import numpy as np
		make_aggregation = {"concatenate": lambda agg_col: sep.join(agg_col), 
					  		"mean": np.mean, "median": np.median, "max": np.max, "min": np.min, "sum": np.sum, "std": np.std, "var": np.var,
					  		"count": lambda agg_col: len(agg_col), "IQR": lambda agg_col: np.percentile(agg_col, 75) - np.percentile(agg_col, 25),
							"PC25": lambda agg_col: np.percentile(agg_col, 25), "PC75": lambda agg_col: np.percentile(agg_col, 75)}
		aggregated_values = []
		for aggregator in aggregation_modes.split(','):
			aggregated_column_copy = aggregated_column.copy()
			if aggregator != "concatenate": aggregated_column_copy = [float(item) for item in aggregated_column_copy]
			agg_value = str(make_aggregation[aggregator](aggregated_column_copy))
			aggregated_values.append(agg_value)
		return "|".join(aggregated_values)

	def records_count(input_table, col_index):
		if len(col_index) == 1: col_index = col_index[0]
		elif len(col_index) > 1: raise ValueError("Only one column can be used to count records")
		
		aggregated_data = defaultdict(lambda: 0 )
		for fields in input_table:
			aggregated_data[fields[col_index]]+=1
			output_list = []
		for k, value in aggregated_data.items():
			output_list.append([k, str(value)])
		return output_list

	def desaggregate_column(input_table, col_index, sep):
		desaggregated_data = []
		for fields in input_table:
			desagg_dict = {col_idx: fields[col_idx].split(sep) for col_idx in col_index}
			for agg_idx in range(len(desagg_dict[col_index[0]])):
				record = fields.copy()
				for col_idx in col_index: record[col_idx] = desagg_dict[col_idx][agg_idx]
				desaggregated_data.append(record)
		return desaggregated_data

	def create_table(indexed_metrics, samples_tag, attributes, metric_names):
		allTags = []
		allTags.extend(attributes)
		allTags.extend(metric_names)
		table_output = []
		corrupted_records = []
		for sample_name, sample_data in indexed_metrics.items():
			record = [sample_name]
			for tag in allTags:
				record.append(sample_data.get(tag))
			if record.count(None) > 0:
				warnings.warn("Record " + sample_name + "is corrupted:\n" + repr(record) + "\n")
				corrupted_records.append(record)
			else:
				table_output.append(record)
		allTags.insert(0, samples_tag)
		table_output.insert(0, allTags) # Add header
		corrupted_records.insert(0, allTags) # Add header
		return table_output, corrupted_records
	
	def name_replaces(tabular_input, sep, cols_to_replace, indexed_file_index, remove_uns=False):
		translated_fields = []
		untranslated_fields = []
		for fields in tabular_input:
			replaced_field = True
			fields = copy.copy(fields) # To avoid modify original data
			for col in cols_to_replace:
				rep_field = indexed_file_index[fields[col]]
				if not rep_field: 
					replaced_field = False
				else:
					fields[col] = rep_field
			if replaced_field:
				translated_fields.append(fields)
			elif not remove_uns:
				translated_fields.append(fields)
			else:
				untranslated_fields.append(fields)
		return translated_fields, untranslated_fields

	def link_table(indexed_linker, tabular_file, drop_line, sep, header=False):
		linked_table = []
		if header: linked_table.append(tabular_file.pop(0)+indexed_linker['header'])
		for fields in tabular_file:
			id = fields[0]
			info_id = indexed_linker.get(id)
			if info_id:
				if type(info_id) == list:
					fields += info_id
				else:
					fields.append(info_id)
				linked_table.append(fields)
			else:
				if not drop_line: linked_table.append(fields) 
		return linked_table

	def merge_tables2mainTab(main_table, table_paths, column_idxs, sep = "\t", fill_character='-', header = False, union= False):
		table_length = len(main_table[0])
		for i, table_path in enumerate(table_paths):
			supp_file = CmdTabs.load_input_data(table_path, sep=sep)
			id_col, data_cols = column_idxs[i]
			supp_file_idx = CmdTabs.index_array(supp_file, id_col, data_cols, header = header)
			for i,row in enumerate(main_table):
				if i == 0 and header: # if we're in first row an is header, search for the header record in the index
					id = 'header'
				else:
					id = row[0]
				supp_data = supp_file_idx.get(id)
				if supp_data:
					if type(supp_data) == list:# list with several row fields
						row.extend(supp_data)
					else:
						row.append(supp_data) # one single field as one single value
				CmdTabs.row_start_fill(row, fill_character, table_length) # When there is no match, fill the gap
			if union: # Add rows present in supp file but not in main_table
				for item in (set(supp_file_idx.keys()) - set([r[0] for r in main_table])):
					new_row = [item]
					CmdTabs.row_start_fill(new_row, fill_character, table_length)
					new_row.extend(supp_file_idx[item])
					main_table.append(new_row)
			table_length += len(supp_file_idx[list(supp_file_idx.keys())[0]])
			
		for row in main_table: CmdTabs.row_end_fill(row, fill_character, table_length) # When there is no match, fill the gap

	def tag_file(input_file, tags, header):
		taged_file = []
		n_row = 0
		for fields in input_file:
			if n_row == 0 and header:
				n_row += 1
				continue
			record = tags.copy()
			record.extend(fields)
			taged_file.append(record)
			n_row += 1
		return taged_file

	def match_pattern(line, all_patterns, search_mode, match_mode, reverse = False	):
		match_pat = False	
		for col, patterns in all_patterns.items():
			is_match = False	
			for pattern in patterns:
				is_match = CmdTabs.expanded_match(line[col], pattern, match_mode)
				if is_match: break	  
			if is_match and search_mode == 's':
				match_pat = False	
				break
			elif not is_match and search_mode == 'c':
				match_pat = True
				break
			elif not is_match:
				match_pat = True
		if reverse: match_pat = not match_pat
		return match_pat

	def expanded_match(string, pattern, match_mode):
		is_match = False
		if pattern in string and match_mode == 'i': is_match = True
		if string == pattern and match_mode == 'c': is_match = True 
		return is_match

	def filter_columns(input_table, cols_to_show, cols_to_match = None, keywords = None, search_mode = None, match_mode = None, reverse = False):
		pattern = CmdTabs.build_pattern(cols_to_match, keywords)
		filtered_table = []
		for line in input_table:
			if not pattern or not CmdTabs.match_pattern(line, pattern, search_mode, match_mode, reverse):
				filtered_table.append(CmdTabs.extract_fields(line, cols_to_show) )
		return filtered_table

	def extract_fields(arr_sub, indexes):
		if indexes == []:
			return arr_sub
		else:
			return [ str(arr_sub[idx]) for idx in indexes] # The str instruction is used to ensure that always we have string data (i.e: when this function is used with excel objecb it could extrac numerical data)

	def merge_files(files, fill_character='-'): #NO TEST
		parent_table = {}
		table_length = 0
		for file_names, file in files.items():
			local_length = 0
			for fields in file:
				id = fields.pop(0) 
				local_length = len(fields)
				CmdTabs.row_start_fill_dict(parent_table, id, fill_character, table_length)
				parent_table[id].extend(fields)				
			table_length += local_length

		parent_table_arr = [] # Fill rows that have not full length (the sum of the row lengths of the all merged tables)
		for id, fields in parent_table.items():
			CmdTabs.row_end_fill(fields, fill_character, table_length)
			record = [id]
			record.extend(fields)
			parent_table_arr.append(record)
		return parent_table_arr

	def row_start_fill_dict(indexed_table, row_id, fill_character, table_length):
		if not indexed_table.get(row_id):
			indexed_table[row_id] = [fill_character] * table_length
		elif len(indexed_table[row_id]) < table_length:
			diference = table_length - len(indexed_table[row_id])
			indexed_table[row_id].extend( [fill_character] * diference)

	def row_start_fill(row, fill_character, table_length):
		n_fields = len(row)
		if n_fields == 0:
			row = [fill_character] * table_length
		elif n_fields < table_length:
			diference = table_length - n_fields
			row.extend( [fill_character] * diference)
		return row

	def row_end_fill(row, fill_character, table_length):
		diference = table_length - len(row)
		if diference > 0: row.extend([fill_character] * diference)

	def merge_and_filter_tables(input_files, options):
		header = []
		filtered_table = []
		if options.get('cols_to_show') == None: options['cols_to_show'] = [*range(0, len(input_files[0][0]), 1)] 
		for filename, file in input_files.items():
			if options.get('header') != None and options['header']:
				if len(header) == 0:
					header = file.pop(0)
				else:
					file.pop(0)
			
			fields = CmdTabs.filter_columns(file, options['cols_to_show'], options['col_filter'], options['keywords'], options['search_mode'], options['match_mode'], options['reverse'])
			filtered_table.extend(fields) 
		if options.get('uniq') != None and options['uniq']: filtered_table = CmdTabs.get_uniq(filtered_table)
		if len(header) > 0:
			header = CmdTabs.extract_fields(header, options['cols_to_show'])
			filtered_table.insert(0, header)
		return filtered_table


	def filter_by_whitelist(table, terms2filter, column2filter, not_exact_match=False):
		if not_exact_match:
			filtered_table = [row for row in table if any([term in row[column2filter] for term in terms2filter])]
		else:
			terms2filter = set(terms2filter)
			terms2filter = {term: True for term in terms2filter}
			filtered_table = [row for row in table if terms2filter.get(row[column2filter])]
		return filtered_table

	def filter_by_blacklist(table, terms2filter, column2filter, not_exact_match=False):
		if not_exact_match:
			filtered_table = [row for row in table if all([term not in row[column2filter] for term in terms2filter])]
		else:
			terms2filter = set(terms2filter)
			terms2filter = {term: True for term in terms2filter}
			filtered_table = [row for row in table if not terms2filter.get(row[column2filter])]
		return filtered_table

	def get_uniq(table):
		return [list(i) for i in set(tuple(i) for i in table)] # list cannot be used by set so we use tuples change back the format

	def extract_columns(table, columns2extract):
		storage = []
		for row in table:
			storage.append(CmdTabs.extract_fields(row, columns2extract))
		return storage
	
	def extract_rows(table, rows2extract):
		if rows2extract == []:
			return table
		else:
			return [ row for idx, row in enumerate(table) if idx in rows2extract ]

	def get_table_from_excel(file, sheet_number):
		import openpyxl
		x = openpyxl.load_workbook(file)
		sheets = x.sheetnames # list excel sheets by name
		ws = x[sheets[sheet_number]] #select sheet by index (so, we select by sheet order)
		sheet = []
		
		for xl_row in ws.iter_rows(min_row=1, min_col=1, max_col=ws.max_column, max_row=ws.max_row, values_only=True):
			row = []
			for val in xl_row:
				if val == None: val = ''
				row.append(val)
			sheet.append(row)

		if CmdTabs.transposed:
			sheet = CmdTabs.transpose(sheet)
			
		return sheet

	def get_groups(a_records, b_records): # inputs are list of string but should be nested lists. This is due to python don't allow to hash list in dicts.
		a_rec = set(a_records)				# For this reason, the last lines convert in lists the strings of the original list to keep the format
		b_rec = set(b_records)				# TODO: See to tranfor to tuples instead to string
		common_set = a_rec & b_rec
		a_only = [[r] for r in a_rec if r not in common_set]
		b_only = [[r] for r in b_rec if r not in common_set]
		common = [[r] for r in common_set]
		return common, a_only, b_only

	def get_subset(common, a_only, b_only, full_a_rec, full_b_rec, keep_subset='c', full=False):
      # As the groups are list with nested list with only one element: [['str1'], ['str2']..] the full mode need to access to 0 element to be use as key in full_X_rec
		if keep_subset == 'c':
			result = common
			if full: result = [full_a_rec[r[0]] + full_b_rec[r[0]] for r in common]
		elif keep_subset == 'a':
			result = a_only
			if full: result = [full_a_rec[r[0]] for r in a_only]
		elif keep_subset == 'b':
			result = b_only
			if full: result = [full_b_rec[r[0]] for r in b_only]
		elif keep_subset == 'ab':
			if full:
				a_only = [full_a_rec[r[0]] for r in a_only]
				b_only = [full_b_rec[r[0]] for r in b_only]
			result = a_only + b_only
		return result

	def transpose(table):
		transposed_table = list(map(list, zip(*table)))
		return transposed_table
	
	def subset_table(table, start_line, number_of_lines, has_header):
		final_line = start_line + number_of_lines 
		if has_header: header = table.pop(0) 
		if final_line > len(table) - 1: final_line = len(table)
		subset_table = table[start_line:final_line]
		if has_header: subset_table.insert(0, header)
		return subset_table 
	
	def transform_to_latex(input_table, header, whole, name):
		latex_table = []
		if header: input_table[0] = [f'\\textbf{{{head_item}}}' for head_item in input_table[0]]
	    
		for idx, row in enumerate(input_table):
			joined_row =  " & ".join(row) + " \\\\ \\hline"		
			if idx == 0: joined_row = "\\hline \n" + joined_row
			latex_table.append(joined_row)
		
		if whole: 
			cols_type = "|"+ "|".join(['l'] * len(input_table[0])) + "|"
			final_table = [[f'\\begin{{table}}[!htbp]\n\\centering\n\\caption{{}}\n\\begin{{tabular}}{{{cols_type}}}']]
			final_table += latex_table
			final_table += [f"\\end{{tabular}}\n\\label{{table:{name}}}\n\\end{{table}}"]
		else:
			final_table = latex_table
		return final_table

	def write_output_data(output_data, output_path=None, sep="\t"):
		open_file = gzip.open if CmdTabs.compressed_output else open
		if CmdTabs.transposed:
			output_data = CmdTabs.transpose(output_data)
			
		if output_path != None:
			with open_file(output_path, 'wt') as out_file:
				for line in output_data:
					out_file.write(sep.join([str(l) for l in line]) + "\n")
		else:
			if CmdTabs.compressed_output:
				columns_joined = []
				for line in output_data:
					columns_joined.append(sep.join([str(l) for l in line]))
				all_joined = "\n".join(columns_joined) + "\n"
				sys.stdout.buffer.write(gzip.compress(bytes(all_joined, 'utf-8')))
			else:
				for line in output_data:
					print(sep.join([str(l) for l in line]))

	def split_by_nFiles(input_table, file_number, output_folder, file_name='table', header = []):
		chunk_size = len(input_table) // file_number
		init_line = 1 if len(header) > 0 else 0
		if len(input_table) % file_number > 0: chunk_size += 1
		chunk_counter = 0
		for idx in range(init_line, len(input_table), chunk_size):
			CmdTabs.write_output_data(header+input_table[idx:idx+chunk_size], os.path.join(output_folder, f"{file_name}_chunk{chunk_counter}"))
			chunk_counter += 1

	def split_by_chunk(input_table, chunk_size, output_folder, file_name='table', header = []):
		init_line = 1 if len(header) > 0 else 0
		chunk_counter = 0
		for idx in range(init_line, len(input_table), chunk_size):
			CmdTabs.write_output_data(header+input_table[idx:idx+chunk_size], os.path.join(output_folder, f"{file_name}_chunk{chunk_counter}"))
			chunk_counter += 1 